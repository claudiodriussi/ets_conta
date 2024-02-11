import os
import sys
import yaml
import openpyxl


class EtsConta:

    def __init__(self, config_file='config.yaml'):
        with open(config_file) as f:
            self.config = yaml.load(f, Loader=yaml.FullLoader)

        # parametri di configurazione
        self.pref_schede = self.config.get("pref_schede", 'SCHEDA_')
        self.pref_eventi = self.config.get("pref_eventi", 'EVENTO_')
        self.pref_anagrafiche = self.config.get("pref_anagrafiche", 'ANAG_')
        self.patrimonio = self.config.get("patrimonio", [])

        # parametri di stile per i fogli excel
        self.title_font = openpyxl.styles.Font(size=15, bold=True)
        self.titrow_fill = openpyxl.styles.PatternFill(
            "solid", start_color='FFFFFFA6')

        # lista dei movimenti
        self.conta = []

    def conti_scan(self, conto):
        """Trova la posizione di un conto nella lista
        """
        for i, c in enumerate(self.lista_conti):
            if c[0] == conto:
                return i

    def read_table(self, filename, tolist=False):
        """Legge una tabella excel con una lista
        """
        xlsx = openpyxl.load_workbook(filename)
        sheet = xlsx.active
        rows = sheet.rows
        table = {}
        table_list = []
        for row in rows:
            # ignora le righe vuote
            if not row:
                continue
            # ignora la riga dei titoli
            if not row[0].value or (row[0].value).upper() == 'CODICE':
                continue
            table[(row[0].value).upper()] = row[1].value
            table_list.append([(row[0].value).upper(), row[1].value])
        return table_list if tolist else table

    def read_files(self, path="."):
        """Legge tutte le tabelle collegate nella cartella che contiene il file
        della primanota.
        """
        def _read_table(key, name, tolist=False):
            name = self.config.get(key, name)
            filename = os.path.join(path, name)
            return self.read_table(filename, tolist=tolist)

        # legge le tabelle delle anagrafiche, dei conti e degli eventi
        self.anag = _read_table("anagrafiche", 'ANAGRAFICHE.xlsx')
        self.eventi = _read_table("eventi", 'EVENTI.xlsx')
        self.conti = _read_table("conti", 'CONTI.xlsx')
        self.lista_conti = _read_table("conti", 'CONTI.xlsx', tolist=True)

        # prepara la directory di destinazione per i documenti e cancella
        # eventuali files vecchi
        self.dest_dir = os.path.join(path, self.config.get("dest_dir", ''))
        os.makedirs(self.dest_dir, exist_ok=True)
        for f in os.listdir(self.dest_dir):
            if not f.endswith(".xlsx"):
                continue
            os.remove(os.path.join(self.dest_dir, f))

    def read_prima(self, filename):
        """ Legge un file di primanota e costruisce la lista delle
        registrazioni.
        Ogni riga di primanota contiene:
        DATA, DESCRIZIONE, CONTO DARE, CONTO AVERE, IMPORTO, EVENTO,
        TAG EVENTO, ANAGRAFICA

        vengono generate due righe, una per il dare ed una per l'avere.
        """

        self.read_files(path=os.path.dirname(filename))
        xlsx = openpyxl.load_workbook(filename)
        sheet = xlsx.active
        rows = sheet.rows

        is_conta = False
        for row in rows:
            # salta le righe vuote
            if not row:
                continue
            # salta le righe dei titoli
            if not row[0].value:
                continue
            if not is_conta and (row[0].value).upper() == 'DATA' and (row[1].value).upper() == 'DESCRIZIONE' and (row[2].value).upper() == 'C.DARE':
                is_conta = True
                continue
            # la lettura dei dati inizia dopo la riga dei titoli
            if not is_conta:
                raise Exception(f"File primanota {filename} non compatibile")
            self.conta.append(ContaRow(row, self, True))
            self.conta.append(ContaRow(row, self, False))

    def write_giornale(self, filename):
        """Scrive il file excel del giornale. Per ogni riga di primanota si
        fanno due righe una per il dare ed una per l'avere
        """

        # riordina i movimenti per data
        self.conta.sort(key=lambda ll: ll.data)

        wb = openpyxl.Workbook()
        ws = wb.active

        ws['B1'] = "GIORNALE DI CONTABILITÀ"
        ws['B1'].font = self.title_font
        ws.append([])
        ws.append(["DATA", "DESCRIZIONE", "CONTO", "NOME",
                  "DARE", "AVERE", "CONTRO", "EVENTO", "TAGS", "ANAGRAFICHE"])
        for cell in ws['A3:J3'][0]:
            cell.fill = self.titrow_fill

        nun_rows = 3
        for row in self.conta:
            rg = [
                row.data,
                row.descriz,
                row.conto,
                row.nome,
                row.impo if row.dare else None,
                - row.impo if not row.dare else None,
                row.contro,
                row.evento,
                row.tag,
                row.anag
            ]
            ws.append(rg)
            nun_rows += 1
            ws[f'A{nun_rows}'].number_format = 'dd/mm/yy'
            ws[f'E{nun_rows}'].number_format = '0.00'
            ws[f'F{nun_rows}'].number_format = '0.00'

        ws[f'D{nun_rows+2}'] = 'Totali: '
        ws[f'E{nun_rows+2}'] = f'=SUM(E4:E{nun_rows})'
        ws[f'E{nun_rows+2}'].number_format = '0.00'
        ws[f'F{nun_rows+2}'] = f'=SUM(F4:F{nun_rows})'
        ws[f'F{nun_rows+2}'].number_format = '0.00'

        # al momento non riesco a settare correttamente la dimensione delle colonne
        # for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        #     openpyxl.worksheet.dimensions.ColumnDimension(
        #         ws, index=i, auto_size=True)
        filename = os.path.join(self.dest_dir, filename)
        wb.save(filename)

    def write_bilancio(self, filename):

        # prepara la lista dei conti
        bilancio = []
        for conto in self.lista_conti:
            bilancio.append([conto[0], conto[1], 0])

        # somma le righe
        for row in self.conta:
            i = self.conti_scan(row.conto)
            bilancio[i][2] += row.impo

        wb = openpyxl.Workbook()
        ws = wb.active

        ws['B1'] = "BILANCIO"
        ws['B1'].font = self.title_font
        ws.append([])
        ws.append(["CONTO", "DESCRIZIONE", "DARE", "AVERE"])
        for cell in ws['A3:D3'][0]:
            cell.fill = self.titrow_fill

        nun_rows = 3
        for row in bilancio:
            if not row[2]:
                continue
            rg = [
                row[0],
                row[1],
                row[2] if row[2] > 0 else None,
                - row[2] if not row[2] > 0 else None,
            ]
            ws.append(rg)
            nun_rows += 1
            ws[f'C{nun_rows}'].number_format = '0.00'
            ws[f'D{nun_rows}'].number_format = '0.00'

        filename = os.path.join(self.dest_dir, filename)
        wb.save(filename)

    def write_schede(self, schede=[]):
        self.conta.sort(key=lambda ll: ll.data)

        conti = {}
        for row in self.conta:
            if schede and row.conto not in schede:
                continue
            if not conti.get(row.conto):
                conti[row.conto] = []
            conti[row.conto].append(row)

        for conto, mov in conti.items():
            self._scheda(conto, mov)

    def _scheda(self, conto, mov):

        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = mov[0].conto
        ws['B1'] = mov[0].nome
        for cell in ws['A1:B1'][0]:
            cell.font = self.title_font

        ws.append([])
        ws.append(["DATA", "DESCRIZIONE", "DARE",
                  "AVERE", "CONTRO", "EVENTO", "TAGS", "ANAGRAFICHE"])
        for cell in ws['A3:H3'][0]:
            cell.fill = self.titrow_fill

        nun_rows = 3
        saldo = 0
        for row in mov:
            rg = [
                row.data,
                row.descriz,
                row.impo if row.dare else None,
                - row.impo if not row.dare else None,
                row.contro,
                row.evento,
                row.tag,
                row.anag
            ]
            saldo += row.impo
            ws.append(rg)
            nun_rows += 1
            ws[f'A{nun_rows}'].number_format = 'dd/mm/yy'
            ws[f'C{nun_rows}'].number_format = '0.00'
            ws[f'D{nun_rows}'].number_format = '0.00'

        ws[f'B{nun_rows+2}'] = 'Totali: '
        ws[f'C{nun_rows+2}'] = f'=SUM(C4:C{nun_rows})'
        ws[f'C{nun_rows+2}'].number_format = '0.00'
        ws[f'D{nun_rows+2}'] = f'=SUM(D4:D{nun_rows})'
        ws[f'D{nun_rows+2}'].number_format = '0.00'
        ws[f'B{nun_rows+3}'] = 'Saldo: '
        col = ['D', 'C'] if saldo < 0 else ['C', 'D']
        ws[f'{col[0]}{nun_rows+3}'] = f'= {col[0]}{nun_rows+2}-{col[1]}{nun_rows+2}'
        ws[f'{col[0]}{nun_rows+3}'].number_format = '0.00'

        filename = os.path.join(self.dest_dir, self.pref_schede+conto+'.xlsx')
        wb.save(filename)

    def write_eventi(self, schede=[]):
        # riordina i movimenti per data
        self.conta.sort(key=lambda row: row.data.isoformat()+row.tag)

        eventi = {}
        for row in self.conta:
            if not row.evento:
                continue
            if schede and row.evento not in schede:
                continue
            if row.conto in self.patrimonio:
                continue
            if not eventi.get(row.evento):
                eventi[row.evento] = []
            eventi[row.evento].append(row)

        for evento, mov in eventi.items():
            self._evento(evento, mov)

    def _evento(self, evento, mov):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = mov[0].evento
        ws['B1'] = self.eventi[mov[0].evento]
        for cell in ws['A1:B1'][0]:
            cell.font = self.title_font

        ws.append([])
        ws.append(["DATA", "DESCRIZIONE", "CONTO",
                  "NOME", "DARE", "AVERE", "CONTRO", "TAGS", "ANAGRAFICHE"])
        for cell in ws['A3:I3'][0]:
            cell.fill = self.titrow_fill

        nun_rows = 3
        saldo = 0
        last_tag = mov[0].tag
        for row in mov:
            if row.tag != last_tag:
                last_tag = row.tag
                nun_rows += 1
                ws.append([])

            rg = [
                row.data,
                row.descriz,
                row.conto,
                row.nome,
                row.impo if row.dare else None,
                - row.impo if not row.dare else None,
                row.contro,
                row.tag,
                row.anag
            ]
            saldo += row.impo
            ws.append(rg)
            nun_rows += 1
            ws[f'A{nun_rows}'].number_format = 'dd/mm/yy'
            ws[f'E{nun_rows}'].number_format = '0.00'
            ws[f'F{nun_rows}'].number_format = '0.00'

        ws[f'D{nun_rows+2}'] = 'Totali: '
        ws[f'E{nun_rows+2}'] = f'=SUM(E4:E{nun_rows})'
        ws[f'E{nun_rows+2}'].number_format = '0.00'
        ws[f'F{nun_rows+2}'] = f'=SUM(F4:F{nun_rows})'
        ws[f'F{nun_rows+2}'].number_format = '0.00'
        ws[f'D{nun_rows+3}'] = 'Saldo: '
        col = ['F', 'E'] if saldo < 0 else ['E', 'F']
        ws[f'{col[0]}{nun_rows+3}'] = f'= {col[0]}{nun_rows+2}-{col[1]}{nun_rows+2}'
        ws[f'{col[0]}{nun_rows+3}'].number_format = '0.00'

        filename = os.path.join(self.dest_dir, self.pref_eventi+evento+'.xlsx')
        wb.save(filename)

    def write_anag(self, schede=[]):
        # riordina i movimenti per data
        self.conta.sort(key=lambda row: row.data.isoformat()+row.tag)

        anag = {}
        for row in self.conta:
            if not row.anag:
                continue
            if schede and row.anag not in schede:
                continue
            if row.conto in self.patrimonio:
                continue
            if not anag.get(row.anag):
                anag[row.anag] = []
            anag[row.anag].append(row)

        for ana, mov in anag.items():
            self._anag(ana, mov)

    def _anag(self, anag, mov):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = mov[0].anag
        ws['B1'] = self.anag[mov[0].anag]
        for cell in ws['A1:B1'][0]:
            cell.font = self.title_font

        ws.append([])
        ws.append(["DATA", "DESCRIZIONE", "CONTO",
                  "NOME", "DARE", "AVERE", "CONTRO", "EVENTO", "TAGS"])
        for cell in ws['A3:I3'][0]:
            cell.fill = self.titrow_fill

        nun_rows = 3
        saldo = 0
        for row in mov:
            rg = [
                row.data,
                row.descriz,
                row.conto,
                row.nome,
                row.impo if row.dare else None,
                - row.impo if not row.dare else None,
                row.contro,
                row.evento,
                row.tag
            ]
            saldo += row.impo
            ws.append(rg)
            nun_rows += 1
            ws[f'A{nun_rows}'].number_format = 'dd/mm/yy'
            ws[f'E{nun_rows}'].number_format = '0.00'
            ws[f'F{nun_rows}'].number_format = '0.00'

        ws[f'D{nun_rows+2}'] = 'Totali: '
        ws[f'E{nun_rows+2}'] = f'=SUM(E4:E{nun_rows})'
        ws[f'E{nun_rows+2}'].number_format = '0.00'
        ws[f'F{nun_rows+2}'] = f'=SUM(F4:F{nun_rows})'
        ws[f'F{nun_rows+2}'].number_format = '0.00'
        ws[f'D{nun_rows+3}'] = 'Saldo: '
        col = ['F', 'E'] if saldo < 0 else ['E', 'F']
        ws[f'{col[0]}{nun_rows+3}'] = f'= {col[0]}{nun_rows+2}-{col[1]}{nun_rows+2}'
        ws[f'{col[0]}{nun_rows+3}'].number_format = '0.00'

        filename = os.path.join(self.dest_dir, self.pref_anagrafiche+anag+'.xlsx')
        wb.save(filename)


class ContaRow:
    def __init__(self, row, conta, dare):
        """Classe d'appoggio per EtsConta.
        Rappresenta una singola riga da cui vengono ricavati tutti i dati
        """
        conto = row[2] if dare else row[3]
        conto = conto.value.upper()
        contro = row[3] if dare else row[2]
        contro = contro.value.upper()
        dsc = conta.conti.get(conto, None)
        if not dsc:
            raise Exception(f"Non trovo il conto {conto}")
        evento = ''
        if row[5].value:
            evento = row[5].value.upper()
            if not conta.eventi.get(evento, None):
                raise Exception(f"Non trovo l'evento {evento}")
        anag = ''
        if row[7].value:
            anag = row[7].value.upper()
            if not conta.anag.get(anag, None):
                raise Exception(f"Non trovo l'anagrafica {anag}")
        self.data = row[0].value
        self.descriz = row[1].value
        self.conto = conto
        self.contro = contro
        self.nome = dsc
        self.dare = dare
        self.impo = row[4].value if dare else -row[4].value
        self.evento = evento
        self.tag = str(row[6].value) if row[6].value else ''
        self.anag = anag


if __name__ == "__main__":
    filename = 'esempio/PRIMANOTA.xlsx' if len(sys.argv) < 2 else sys.argv[1]
    conta = EtsConta()
    conta.read_prima(filename)
    conta.write_giornale('GIORNALE.xlsx')
    conta.write_schede()
    conta.write_eventi()
    conta.write_anag()
    conta.write_bilancio('BILANCIO.xlsx')
